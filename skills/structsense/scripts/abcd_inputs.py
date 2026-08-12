"""Resolve whatever the user points at into a list of local PDFs.

One argument, no mode flags. What it accepts:

  * a **PDF / TXT / MD file**            -> that one paper
  * a **directory**                      -> every paper under it, recursively
  * a **CSV / TSV / XLSX of DOIs**       -> fetch the open-access PDF for each
  * a **.txt list of DOIs or URLs**      -> same
  * a **DOI or URL** given directly      -> same

Bulk is not a mode; it is what happens when the input contains more than one
paper. `resolve()` returns one list either way, so the caller has no branch.

Fetching is **open access only**. Unpaywall (with an email, per their terms) then
OpenAlex then Semantic Scholar, all free, no credentials. Paywalled papers are
reported as unresolved with the reason — this module will not touch a publisher
paywall or an institutional proxy. If you have entitled access, download those
PDFs yourself and point at the directory; the sibling `synthscholar` skill has an
EZproxy path for that case.

Every fetch records provenance: which service answered, the URL, the license where
known, and a sha256 of the bytes.
"""
from __future__ import annotations

# Run either way: `python -m scripts.<mod>` from the skill directory, or
# `python /abs/path/to/scripts/<mod>.py` from anywhere. Without this, running the
# file directly fails with ModuleNotFoundError: scripts — which forces callers to
# cd into the skill first, for no reason.
if __package__ in (None, ""):  # executed as a file, not as part of the package
    import sys as _sys
    from pathlib import Path as _Path
    _sys.path.insert(0, str(_Path(__file__).resolve().parent.parent))

import csv
import hashlib
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

PAPER_SUFFIXES = (".pdf", ".txt", ".md")
TABLE_SUFFIXES = (".csv", ".tsv", ".tab", ".xlsx")
DOI_RE = re.compile(r"10\.\d{4,9}/[^\s\"'<>,;)\]]+", re.I)

UNPAYWALL = "https://api.unpaywall.org/v2/{doi}?email={email}"
OPENALEX = "https://api.openalex.org/works/doi:{doi}"
SEMANTIC = "https://api.semanticscholar.org/graph/v1/paper/DOI:{doi}?fields=openAccessPdf,title"

# DOI columns, in preference order. Titles are kept for reporting but are not used
# to guess a DOI — resolving a paper by fuzzy title match is how the wrong paper
# ends up in a corpus.
DOI_COLUMNS = ("doi", "dois", "doi_url", "article_doi", "paper_doi", "identifier")
TITLE_COLUMNS = ("title", "paper_title", "article_title", "name")


@dataclass
class Paper:
    """One resolved (or unresolved) input paper."""

    path: Optional[Path] = None
    doi: Optional[str] = None
    title: Optional[str] = None
    origin: str = "local"          # local | fetched
    provenance: Dict[str, Any] = field(default_factory=dict)
    error: Optional[str] = None

    @property
    def ok(self) -> bool:
        return self.path is not None and self.path.is_file()


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #

def normalize_doi(raw: str) -> Optional[str]:
    """Extract a bare DOI from a string, URL, or 'doi:' form."""
    if not raw:
        return None
    m = DOI_RE.search(str(raw).strip())
    if not m:
        return None
    return m.group(0).rstrip(".,;)").lower()


def _slug_doi(doi: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", doi.lower()).strip("_")


def _get_json(url: str, timeout: int = 30) -> Optional[Any]:
    try:
        req = urllib.request.Request(
            url, headers={"Accept": "application/json",
                          "User-Agent": "structsense-abcd/0.5 (+skill)"}
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode("utf-8", "replace"))
    except Exception:
        return None


def _download(url: str, dest: Path, timeout: int = 90) -> Tuple[bool, str]:
    try:
        req = urllib.request.Request(
            url, headers={"User-Agent": "structsense-abcd/0.5 (+skill)"}
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            ctype = (resp.headers.get("Content-Type") or "").lower()
            blob = resp.read()
    except Exception as exc:
        return False, f"download_failed: {exc}"
    # A paywall or captcha usually answers 200 with HTML. Check the magic bytes
    # rather than the content type, which lies often enough to matter.
    if not blob.startswith(b"%PDF"):
        head = blob[:200].decode("utf-8", "replace").strip().replace("\n", " ")
        return False, (f"not_a_pdf (content-type={ctype or 'unknown'}, "
                       f"starts with {head[:60]!r})")
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(blob)
    return True, hashlib.sha256(blob).hexdigest()


# --------------------------------------------------------------------------- #
# open-access resolution
# --------------------------------------------------------------------------- #

def oa_pdf_url(doi: str, *, email: Optional[str] = None) -> Tuple[Optional[str], dict]:
    """Find a free full-text PDF URL for `doi`. Returns (url, provenance)."""
    tried: List[dict] = []

    if email:
        data = _get_json(UNPAYWALL.format(doi=urllib.parse.quote(doi), email=email))
        loc = (data or {}).get("best_oa_location") or {}
        url = loc.get("url_for_pdf") or loc.get("url")
        tried.append({"service": "unpaywall", "hit": bool(url)})
        if url:
            return url, {"service": "unpaywall", "url": url,
                         "license": loc.get("license"),
                         "version": loc.get("version"), "tried": tried}
    else:
        tried.append({"service": "unpaywall", "hit": False,
                      "skipped": "no email (their terms require one)"})

    data = _get_json(OPENALEX.format(doi=urllib.parse.quote(doi)))
    if data:
        loc = data.get("best_oa_location") or data.get("primary_location") or {}
        url = loc.get("pdf_url")
        tried.append({"service": "openalex", "hit": bool(url)})
        if url:
            return url, {"service": "openalex", "url": url,
                         "license": loc.get("license"),
                         "version": loc.get("version"), "tried": tried}
    else:
        tried.append({"service": "openalex", "hit": False})

    data = _get_json(SEMANTIC.format(doi=urllib.parse.quote(doi)))
    url = ((data or {}).get("openAccessPdf") or {}).get("url")
    tried.append({"service": "semantic_scholar", "hit": bool(url)})
    if url:
        return url, {"service": "semantic_scholar", "url": url, "tried": tried}

    return None, {"service": None, "tried": tried}


def fetch_doi(doi: str, out_dir: Path, *, email: Optional[str] = None,
              delay: float = 0.34) -> Paper:
    """Fetch one DOI's open-access PDF into `out_dir` (cached by DOI slug)."""
    dest = out_dir / f"{_slug_doi(doi)}.pdf"
    if dest.is_file() and dest.stat().st_size > 0:
        return Paper(path=dest, doi=doi, origin="cached",
                     provenance={"cached": True, "path": str(dest)})

    url, prov = oa_pdf_url(doi, email=email)
    time.sleep(delay)                       # be a polite client of free APIs
    if not url:
        return Paper(doi=doi, origin="fetched", provenance=prov,
                     error="no_open_access_pdf_found")

    ok, detail = _download(url, dest)
    if not ok:
        return Paper(doi=doi, origin="fetched",
                     provenance={**prov, "download_error": detail},
                     error=detail)
    return Paper(path=dest, doi=doi, origin="fetched",
                 provenance={**prov, "sha256": detail,
                             "retrieved_at": time.strftime(
                                 "%Y-%m-%dT%H:%M:%SZ", time.gmtime())})


# --------------------------------------------------------------------------- #
# table / list parsing
# --------------------------------------------------------------------------- #

def _pick(headers: Sequence[str], wanted: Sequence[str]) -> Optional[str]:
    norm = {re.sub(r"[^a-z0-9]+", "_", (h or "").strip().lower()).strip("_"): h
            for h in headers}
    for w in wanted:
        if w in norm:
            return norm[w]
    return None


def read_table(path: Path) -> List[Tuple[Optional[str], Optional[str]]]:
    """Read (doi, title) pairs from a CSV/TSV/XLSX."""
    rows: List[Dict[str, Any]] = []
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as exc:
            raise RuntimeError("reading .xlsx needs openpyxl (pip install openpyxl)") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            hdr = [str(h) if h is not None else "" for h in (next(it, None) or [])]
            for r in it:
                rows.append({hdr[i]: r[i] for i in range(min(len(hdr), len(r)))})
        finally:
            wb.close()
        headers = hdr
    else:
        text = path.read_text(encoding="utf-8-sig", errors="replace")
        try:
            delim = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|").delimiter
        except csv.Error:
            delim = "\t" if path.suffix.lower() in (".tsv", ".tab") else ","
        reader = csv.DictReader(text.splitlines(), delimiter=delim)
        headers = list(reader.fieldnames or [])
        rows = [dict(r) for r in reader]

    doi_col = _pick(headers, DOI_COLUMNS)
    title_col = _pick(headers, TITLE_COLUMNS)

    out: List[Tuple[Optional[str], Optional[str]]] = []
    for r in rows:
        doi = normalize_doi(str(r.get(doi_col) or "")) if doi_col else None
        if not doi:
            # No DOI column, or an empty cell — scan the whole row, since exports
            # often bury the DOI in a URL column.
            doi = normalize_doi(" ".join(str(v) for v in r.values() if v))
        title = str(r.get(title_col) or "").strip() if title_col else None
        if doi or title:
            out.append((doi, title or None))
    return out


def _read_list(path: Path) -> List[Tuple[Optional[str], Optional[str]]]:
    out = []
    for line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        doi = normalize_doi(line)
        if doi:
            out.append((doi, None))
    return out


# --------------------------------------------------------------------------- #
# the one entry point
# --------------------------------------------------------------------------- #

def resolve(target: str | Path, *, download_dir: Optional[Path] = None,
            email: Optional[str] = None, limit: Optional[int] = None,
            progress: bool = True) -> Tuple[List[Paper], dict]:
    """Turn one argument into a list of papers. Single vs bulk is inferred."""
    email = email or os.getenv("STRUCTSENSE_EMAIL") or os.getenv("SYNTHSCHOLAR_EMAIL")
    raw = str(target)
    p = Path(raw).expanduser()
    papers: List[Paper] = []
    kind: str

    if p.is_dir():
        kind = "directory"
        files = sorted(f for f in p.rglob("*") if f.suffix.lower() in PAPER_SUFFIXES)
        # Drop extracted-text sidecars. `--prepare` writes <stem>.txt next to each
        # PDF, and counting both would process the same paper twice — and, because
        # outputs are named from the stem, have the second run overwrite the first.
        pdf_stems = {f.with_suffix("").as_posix() for f in files
                     if f.suffix.lower() == ".pdf"}
        files = [f for f in files
                 if f.suffix.lower() == ".pdf"
                 or f.with_suffix("").as_posix() not in pdf_stems]
        papers = [Paper(path=f, title=f.stem) for f in files]

    elif p.is_file() and p.suffix.lower() in PAPER_SUFFIXES:
        kind = "single_file"
        # A .txt of DOIs is a list, not a paper — decide by looking inside.
        if p.suffix.lower() == ".txt":
            entries = _read_list(p)
            if entries:
                kind = "doi_list"
                papers = _fetch_all(entries, p, download_dir, email, limit, progress)
        if not papers:
            papers = [Paper(path=p, title=p.stem)]

    elif p.is_file() and p.suffix.lower() in TABLE_SUFFIXES:
        kind = "doi_table"
        entries = read_table(p)
        papers = _fetch_all(entries, p, download_dir, email, limit, progress)

    elif normalize_doi(raw):
        kind = "doi"
        doi = normalize_doi(raw)
        out = download_dir or Path.cwd() / "abcd_pdfs"
        papers = [fetch_doi(doi, out, email=email)] if doi else []

    elif p.exists():
        raise ValueError(
            f"{p} is not a paper ({'/'.join(PAPER_SUFFIXES)}), a DOI table "
            f"({'/'.join(TABLE_SUFFIXES)}), or a directory"
        )
    else:
        raise FileNotFoundError(f"{raw} does not exist and is not a DOI")

    seen_total = len(papers)
    papers, duplicates = _drop_duplicate_files(papers)
    resolved = [x for x in papers if x.ok]
    summary = {
        "input": raw,
        "detected_as": kind,
        "papers_total": len(papers),
        "files_seen": seen_total,
        "papers_resolved": len(resolved),
        "papers_unresolved": len(papers) - len(resolved),
        "open_access_email_used": bool(email),
        "duplicates_dropped": len(duplicates),
        "duplicates": duplicates[:50],
        "unresolved": [
            {"doi": x.doi, "title": x.title, "reason": x.error}
            for x in papers if not x.ok
        ][:200],
    }
    return resolved, summary


def _drop_duplicate_files(papers: List[Paper]) -> Tuple[List[Paper], List[dict]]:
    """Keep one paper per distinct file content.

    A download directory routinely holds `Aaron-2025-….pdf` and
    `Aaron-2025-…(1).pdf` — byte-identical copies from two download attempts.
    Processing both extracts the same paper twice, and because the synthesis counts
    by paper, a duplicate silently doubles one study's weight in every consensus and
    turns "two papers agree" into a fact about one paper. Content hashing catches
    the "(1)" copies, renamed copies and copies in different subdirectories alike;
    an unreadable file is kept rather than guessed about.
    """
    seen: Dict[str, Paper] = {}
    kept: List[Paper] = []
    dupes: List[dict] = []
    for pap in papers:
        path = pap.path
        if not path or not Path(path).is_file():
            kept.append(pap)
            continue
        try:
            digest = _sha256(Path(path))
        except OSError:
            kept.append(pap)
            continue
        first = seen.get(digest)
        if first is None:
            seen[digest] = pap
            kept.append(pap)
            continue
        # Keep the cleaner filename. Output files are named from the stem, so
        # keeping `paper(1).pdf` over `paper.pdf` produces `paper(1)_abcd.json` and
        # stops a `paper.payload.json` from ever being found.
        if _copy_rank(Path(path)) < _copy_rank(Path(str(first.path))):
            kept[kept.index(first)] = pap
            seen[digest] = pap
            pap, first = first, pap
            path = Path(str(pap.path))
        dupes.append({"dropped": str(path), "kept": str(first.path),
                      "sha256": digest[:16], "reason": "identical file content"})
        print(f"  skipping duplicate {Path(path).name} — identical to "
              f"{Path(str(first.path)).name}", file=sys.stderr)
    return kept, dupes


_COPY_MARKER = re.compile(r"[ _-]*\((\d+)\)$|[ _-]copy(?:\s*\d+)?$", re.I)


def _copy_rank(path: Path) -> tuple:
    """Lower is preferred: no "(1)"/"copy" marker, then the shorter name."""
    stem = path.stem
    return (1 if _COPY_MARKER.search(stem) else 0, len(stem), stem)


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1 << 20), b""):
            h.update(block)
    return h.hexdigest()


def _fetch_all(entries: List[Tuple[Optional[str], Optional[str]]], src: Path,
               download_dir: Optional[Path], email: Optional[str],
               limit: Optional[int], progress: bool) -> List[Paper]:
    out_dir = download_dir or (src.parent / "abcd_pdfs")
    todo = entries[:limit] if limit else entries
    papers: List[Paper] = []
    for i, (doi, title) in enumerate(todo, 1):
        if not doi:
            papers.append(Paper(title=title, error="no_doi_in_row"))
            continue
        pap = fetch_doi(doi, out_dir, email=email)
        pap.title = pap.title or title
        papers.append(pap)
        if progress:
            state = "cached" if pap.origin == "cached" else (
                "ok" if pap.ok else f"MISS ({pap.error})")
            print(f"  [{i}/{len(todo)}] {doi} -> {state}", file=sys.stderr)
    return papers


def _cli(argv: Optional[List[str]] = None) -> int:
    import argparse

    ap = argparse.ArgumentParser(description=(__doc__ or "").splitlines()[0])
    ap.add_argument("input", help="PDF, directory, DOI table, DOI list, or a DOI")
    ap.add_argument("--download-dir", type=Path)
    ap.add_argument("--email", help="for Unpaywall (their terms require one)")
    ap.add_argument("--limit", type=int)
    a = ap.parse_args(argv)

    papers, summary = resolve(a.input, download_dir=a.download_dir, email=a.email,
                              limit=a.limit)
    print(json.dumps(summary, indent=1))
    for p in papers:
        print(f"{p.origin:7} {p.path}")
    return 0 if papers else 1


if __name__ == "__main__":
    raise SystemExit(_cli())
