"""ABCD / HBCD variable dictionary — load a release snapshot, then resolve names.

The authoritative variable list is the NBDC data dictionary (`lst_dds`), keyed by
study (`abcd` | `hbcd`) and release (`"5.1"`, `"6.0"`, ...). It ships in the
`NBDCtoolsData` R data package; `nbdctools` (PyPI) downloads and reads it without
R. This module wraps that into three operations the ABCD extraction mode needs:

  1. `build_snapshot()` — materialise one study+release as a flat JSON snapshot on
     disk, with provenance (source, retrieval time, sha256, row count).
  2. `load_snapshot()` / `load_all_snapshots()` — read snapshots back.
  3. `Dictionary.resolve()` — decide whether a string from a paper IS a real
     variable in that release, and how we know.

Why a snapshot rather than querying live: a claim like "this paper uses
`nihtbx_flanker_uncorrected`" is only checkable against a *stated* release. The
snapshot pins it, so a re-run months later reproduces the same verdict, and the
provenance block records exactly which dictionary was consulted.

ACQUISITION IS TOOL-ONLY. There is no "the model knows ABCD variable names"
path — a variable that cannot be found in a real dictionary is reported as
unverified, never silently accepted. This mirrors the skill's concept-mapping
rule (SKILL.md hard rule 15).

    # once per release (needs: pip install nbdctools)
    python -m scripts.abcd_dictionary build --study abcd --release 6.1
    python -m scripts.abcd_dictionary build --study abcd --release latest --all-releases

    # or from a dictionary you exported yourself (R, or the NBDC portal)
    python -m scripts.abcd_dictionary build --study abcd --release 6.1 \
        --from-csv my_dd_export.csv

    python -m scripts.abcd_dictionary lookup nihtbx_flanker_uncorrected
    python -m scripts.abcd_dictionary info
"""
from __future__ import annotations

# Run either way: `python -m scripts.<mod>` from the skill directory, or
# `python /abs/path/to/scripts/<mod>.py` from anywhere. Without this, running the
# file directly fails with ModuleNotFoundError: scripts — which forces callers to
# cd into the skill first, for no reason.
if __package__ in (None, ""):  # executed as a file, not as part of the package
    import sys as _sys
    from pathlib import Path as _Path
    _sys.path.insert(0, str(_Path(__file__).resolve().parent.parent))

import argparse
import hashlib
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

# Where snapshots live. Override with STRUCTSENSE_ABCD_DIR.
DEFAULT_DIR = Path(
    os.getenv("STRUCTSENSE_ABCD_DIR")
    or (Path.home() / ".cache" / "structsense" / "abcd")
)

# Dictionaries bundled WITH the skill, so it is self-contained: no file in
# ~/Downloads, no network, no R. Minimal columns + gzip keeps ABCD 6.1 at ~1.9MB
# and the NDA-era set at ~1.1MB, against 76MB and 42MB uncompressed.
BUNDLED_DIR = Path(__file__).resolve().parent.parent / "data" / "dictionaries"

# Columns a bundled snapshot keeps. Enough to verify a mention, resolve every
# alternate naming, and report table/domain — everything the ABCD mode reads.
MINIMAL_COLUMNS = (
    "name", "label", "table_name", "table_nda", "domain", "sub_domain",
    "name_nda", "name_deap", "name_short",
)

# r-universe endpoint nbdctools itself uses for the dictionary bundle.
LST_DDS_URL = "https://nbdc-datahub.r-universe.dev/NBDCtoolsData/data/lst_dds/rds"

# The ABCD documentation site publishes release notes per data release. It is not a
# machine-readable dictionary, but it IS a public, citable list of which releases
# exist — useful for two things: catching a paper that states a release that never
# shipped, and giving each snapshot a citation URL. ABCD only; HBCD documents
# separately.
ABCD_RELEASE_NOTES_INDEX = (
    "https://docs.abcdstudy.org/latest/documentation/release_notes/"
)

# Columns we keep. The dictionary has many more; these are the ones that let us
# verify a mention and describe it back to the user.
KEEP_COLUMNS = (
    "name",
    "label",
    "description",
    "table_name",
    "table_label",
    # NDA structure name. Kept because the NBDC catalog carries both namings and a
    # paper may cite either.
    "table_nda",
    "nda_or_nbdc_table",
    "domain",
    "nbdc_domain",
    "sub_domain",
    "nbdc_sub_domain",
    "source",
    "metric",
    "atlas",
    "type_data",
    "type_level",
    "type_var",
    "unit",
    "level_range",
    # Alternate variable names (see NAME_ALIAS_COLUMNS). ABCD 6.x renamed variables
    # wholesale — `nc_y_flnkr_adm___1` in 6.1 is `neurocog_2_flanker___1` in NDA and
    # `neurocog_2_flanker` in DEAP — so dropping these would make every paper that
    # cites an NDA or DEAP name look unverifiable.
    "name_nda",
    "name_deap",
    "name_redcap",
    "name_redcap_exp",
    "name_short",
    "name_stata",
    "url_table",
    "url_docs_score",
)

# The dictionary's column naming differs between the NBDCtools bundle and a
# portal/CSV export (`table_name` vs `nda_or_nbdc_table`, `domain` vs
# `nbdc_domain`). Output always uses the right-hand canonical name, filled from
# whichever column the loaded snapshot actually has, so downstream consumers get
# one stable shape regardless of where the dictionary came from.
FIELD_ALIASES = {
    # `table_nda` is the NDA structure name in the NBDC variable catalog; prefer an
    # explicit nda_or_nbdc_table when an export supplies one, else the catalog's
    # NDA table, else the NBDC table.
    "nda_or_nbdc_table": ("nda_or_nbdc_table", "table_nda", "table_name"),
    "nbdc_domain": ("nbdc_domain", "domain"),
    "nbdc_sub_domain": ("nbdc_sub_domain", "sub_domain"),
}

# Columns of the NBDC variable catalog that hold ALTERNATE names for the same
# variable. Papers cite whichever naming their pipeline used — NDA element names
# and DEAP names are both common in methods sections — so each is indexed as a way
# in, and the match method records which naming the paper used.
NAME_ALIAS_COLUMNS = {
    "name_nda": "nda_name",
    "name_deap": "deap_name",
    "name_redcap": "redcap_name",
    "name_redcap_exp": "redcap_name",
    "name_short": "short_name",
    "name_stata": "stata_name",
}


def canonical_field(row: dict, field: str):
    """First non-empty value among `field`'s known source columns."""
    for col in FIELD_ALIASES.get(field, (field,)):
        val = row.get(col)
        if val not in (None, ""):
            return val
    return None

_WS = re.compile(r"\s+")
# ABCD/HBCD variable names are snake_case ASCII: nihtbx_flanker_uncorrected,
# smri_vol_cdk_banksstslh, ab_g_dyn__visit_type. Used to tell a plausible
# variable token from an English phrase.
VARNAME_RE = re.compile(r"^[a-z][a-z0-9]*(?:_{1,2}[a-z0-9]+)+$")


class DictionaryError(RuntimeError):
    """Raised when no real dictionary can be obtained — never fall back to guessing."""


# --------------------------------------------------------------------------- #
# normalisation
# --------------------------------------------------------------------------- #

def norm_name(s: str) -> str:
    """Canonical form for variable-name comparison.

    Lowercase and strip surrounding punctuation/whitespace. Internal underscores
    are significant (`ab_g_dyn__visit_type` differs from `ab_g_dyn_visit_type`),
    so they are NOT collapsed.
    """
    return (s or "").strip().strip("`'\"“”‘’.,;:()[]{}").lower()


def norm_text(s: str) -> str:
    """Loose form for label/description comparison: lowercase, collapse spaces."""
    return _WS.sub(" ", (s or "").strip().lower())


def looks_like_variable_name(s: str) -> bool:
    """True if `s` has the shape of an ABCD/HBCD variable name."""
    return bool(VARNAME_RE.match(norm_name(s)))


# --------------------------------------------------------------------------- #
# snapshot building
# --------------------------------------------------------------------------- #

def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1 << 20), b""):
            h.update(block)
    return h.hexdigest()


def _rows_from_nbdctools(study: str, release: str, *, rds_path: Optional[Path],
                         progress: bool) -> Tuple[List[dict], dict]:
    """Load `lst_dds` via nbdctools and return (rows, provenance) for one release."""
    try:
        from nbdctools import download_metadata, load_metadata  # type: ignore
    except ImportError as exc:  # pragma: no cover - depends on user's env
        raise DictionaryError(
            "nbdctools is not installed, so the ABCD/HBCD dictionary cannot be "
            "fetched. Either `pip install nbdctools`, or pass --from-csv with a "
            "dictionary you exported yourself (e.g. in R: "
            "write.csv(NBDCtools::get_dd('abcd', '6.1'), 'dd.csv')). "
            "This step is not optional: variable claims are verified against a "
            "real dictionary or reported as unverified."
        ) from exc

    if rds_path is not None:
        path = Path(rds_path)
        if not path.is_file():
            raise DictionaryError(f"--from-rds {path} does not exist")
        source = str(path)
    else:
        try:
            path = Path(download_metadata(type="dds", progress=progress))
        except Exception as exc:
            raise DictionaryError(
                f"Could not download the dictionary bundle ({exc}). "
                f"Fetch {LST_DDS_URL} in a browser and re-run with "
                "--from-rds <downloaded lst_dds.rds>, or use --from-csv."
            ) from exc
        source = LST_DDS_URL

    dds = load_metadata(path, progress=progress)
    if not isinstance(dds, dict) or study not in dds:
        available = ", ".join(sorted(dds)) if isinstance(dds, dict) else "?"
        raise DictionaryError(f"study {study!r} not in the bundle (have: {available})")
    by_release = dds[study]
    releases = sorted(by_release, key=_release_sort_key)
    if release in ("latest", ""):
        release = releases[-1]
    if release not in by_release:
        raise DictionaryError(
            f"release {release!r} not in the bundle for {study!r} "
            f"(have: {', '.join(releases)})"
        )

    rows = _to_rows(by_release[release])
    prov = {
        "source": source,
        "source_sha256": _sha256_file(path) if path.is_file() else None,
        "nbdctools_data_releases_available": releases,
    }
    return rows, {**prov, "resolved_release": release}


def _release_sort_key(rel: str) -> tuple:
    """Sort '5.1' < '6.0' < '6.1' numerically, tolerating odd labels."""
    parts = re.findall(r"\d+", str(rel))
    return (tuple(int(p) for p in parts) if parts else (0,), str(rel))


def _to_rows(dd: Any) -> List[dict]:
    """Coerce a polars/pandas/dict dictionary table into a list of plain dicts."""
    if hasattr(dd, "to_dicts"):          # polars
        rows = dd.to_dicts()
    elif hasattr(dd, "to_dict"):         # pandas
        rows = dd.to_dict(orient="records")
    elif isinstance(dd, dict):           # column-oriented dict
        keys = list(dd)
        n = len(dd[keys[0]]) if keys else 0
        rows = [{k: dd[k][i] for k in keys} for i in range(n)]
    elif isinstance(dd, list):
        rows = list(dd)
    else:
        raise DictionaryError(f"unrecognised dictionary object: {type(dd).__name__}")
    return [_keep(r) for r in rows if (r.get("name") or "").strip()]


def _keep(row: dict) -> dict:
    out = {}
    for k in KEEP_COLUMNS:
        v = row.get(k)
        if v is None:
            continue
        out[k] = v if isinstance(v, (str, int, float, bool)) else str(v)
    return out


# Header aliases for a hand-supplied export. Different sources name the same
# columns differently — an R `get_dd()` dump, a DEAP (abcd.deapscience.com)
# create-dataset variable export, and an NDA data-dictionary download all differ —
# so accept the common spellings instead of making the user rename columns. Keys
# are canonical; values are lowercased/underscored headers to look for, in order.
CSV_HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "name": ("name", "element_name", "variable_name", "variable", "element",
             "item_name", "field_name", "short_name"),
    "label": ("label", "element_description", "variable_label", "item_label",
              "title", "element_label"),
    "description": ("description", "notes", "element_notes", "definition",
                    "long_description"),
    "nda_or_nbdc_table": ("nda_or_nbdc_table", "table_name", "table", "structure",
                          "nda_structure", "instrument", "form"),
    "nbdc_domain": ("nbdc_domain", "domain", "category", "construct_domain"),
    "nbdc_sub_domain": ("nbdc_sub_domain", "sub_domain", "subdomain", "subcategory"),
    "type_data": ("type_data", "data_type", "type", "value_type"),
    "unit": ("unit", "units"),
    "level_range": ("level_range", "value_range", "range", "notes_values"),
}


def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (h or "").strip().lower()).strip("_")


def _rows_from_csv(path: Path) -> Tuple[List[dict], dict]:
    """Load a user-supplied dictionary export (CSV or TSV, flexible headers)."""
    import csv

    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = raw[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delim = dialect.delimiter
    except csv.Error:
        delim = "\t" if path.suffix.lower() in (".tsv", ".tab") else ","

    reader = csv.DictReader(raw.splitlines(), delimiter=delim)
    headers = [h for h in (reader.fieldnames or []) if h]
    if not headers:
        raise DictionaryError(f"{path} has no header row")

    # canonical field -> the actual header we will read it from
    norm = {_norm_header(h): h for h in headers}
    mapping: Dict[str, str] = {}
    for canonical, candidates in CSV_HEADER_ALIASES.items():
        for cand in candidates:
            if cand in norm:
                mapping[canonical] = norm[cand]
                break

    if "name" not in mapping:
        raise DictionaryError(
            f"{path} has no recognisable variable-name column. Saw: "
            f"{', '.join(headers[:12])}"
            f"{'…' if len(headers) > 12 else ''}. Accepted spellings: "
            f"{', '.join(CSV_HEADER_ALIASES['name'])}. Rename the column, or "
            "export from R with write.csv(NBDCtools::get_dd('abcd', '6.1'), …)."
        )

    rows: List[dict] = []
    for rec in reader:
        translated = {
            canonical: (rec.get(header) or "").strip()
            for canonical, header in mapping.items()
        }
        if translated.get("name"):
            rows.append(_keep(translated))

    if not rows:
        raise DictionaryError(
            f"{path} produced no rows with a non-empty "
            f"{mapping['name']!r} value."
        )
    return rows, {
        "source": str(path),
        "source_sha256": _sha256_file(path),
        "source_delimiter": "tab" if delim == "\t" else delim,
        # Record the header translation: a reader must be able to see which of the
        # export's columns became `name`, `nbdc_domain`, and so on.
        "csv_header_mapping": mapping,
        "csv_headers_ignored": sorted(
            h for h in headers if h not in set(mapping.values())
        ),
        "nbdctools_data_releases_available": None,
    }


# NDA's data-dictionary API. Structure DEFINITIONS are public (the data behind them
# is not), which makes it the source for the pre-6.0 era: ABCD releases 4.x and 5.x
# were distributed through NDA, so those are the names older papers cite
# (`nihtbx_flanker_uncorrected` lives in structure `abcd_tbss01`). The NBDC catalog
# workbook only goes back to 6.0.
NDA_DICT_API = "https://nda.nih.gov/api/datadictionary/datastructure"


def _http_json(url: str, *, timeout: int = 60) -> Any:
    import urllib.request

    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8", "replace"))


def _rows_from_nda(study: str, *, cache_dir: Path, prefix: Optional[str] = None,
                   delay: float = 0.15, progress: bool = False,
                   refresh: bool = False) -> Tuple[List[dict], dict]:
    """Build a dictionary from NDA structure definitions.

    One request lists every structure; one more per matching structure lists its
    data elements. Responses are cached on disk, so a rebuild is free and a run
    interrupted halfway resumes without re-fetching. Requests are throttled — this
    is a public API being asked for a few hundred documents.
    """
    ndir = cache_dir / "nda"
    ndir.mkdir(parents=True, exist_ok=True)
    index_path = ndir / "datastructure_index.json"

    if index_path.is_file() and not refresh:
        index = json.loads(index_path.read_text())
    else:
        try:
            index = _http_json(NDA_DICT_API)
        except Exception as exc:
            raise DictionaryError(
                f"could not reach the NDA data dictionary ({exc}). "
                f"Fetch {NDA_DICT_API} yourself and place it at {index_path}."
            ) from exc
        index_path.write_text(json.dumps(index))

    structures = index if isinstance(index, list) else index.get("datastructure", [])
    pref = (prefix or f"{study}_").lower()
    matching = [s for s in structures
                if str(s.get("shortName", "")).lower().startswith(pref)]
    if not matching:
        raise DictionaryError(
            f"no NDA structures start with {pref!r} (the index has "
            f"{len(structures)} structures). Pass --nda-prefix to widen."
        )

    rows: List[dict] = []
    fetched = failed = 0
    for i, meta in enumerate(matching, 1):
        short = str(meta["shortName"])
        cache = ndir / f"{short}.json"
        if cache.is_file() and not refresh:
            try:
                doc = json.loads(cache.read_text())
            except Exception:
                doc = None
        else:
            doc = None
        if doc is None:
            try:
                doc = _http_json(f"{NDA_DICT_API}/{short}")
                cache.write_text(json.dumps(doc))
                fetched += 1
                time.sleep(delay)
            except Exception as exc:
                failed += 1
                print(f"  warning: {short}: {exc}", file=sys.stderr)
                continue

        title = str(doc.get("title") or meta.get("title") or "").strip()
        cats = meta.get("categories") or doc.get("categories") or []
        domain = ", ".join(str(c) for c in cats) if isinstance(cats, list) else str(cats)
        for el in doc.get("dataElements") or []:
            name = str(el.get("name") or "").strip()
            if not name:
                continue
            row = {
                "name": name,
                "label": str(el.get("description") or "").strip(),
                "description": str(el.get("notes") or "").strip(),
                # For an NDA-sourced dictionary the structure IS the NDA table.
                "table_nda": short,
                "table_name": short,
                "table_label": title,
                "domain": domain,
                "type_data": str(el.get("type") or el.get("dataType") or "").strip(),
                "unit": str(el.get("size") or "").strip(),
            }
            aliases = el.get("aliases")
            if isinstance(aliases, str):
                aliases = aliases.strip()
                if aliases and aliases not in ("[]", "None"):
                    row["name_nda"] = aliases.strip("[]'\" ")
            elif isinstance(aliases, list) and aliases:
                row["name_nda"] = str(aliases[0])
            rows.append(_keep(row))
        if progress and i % 25 == 0:
            print(f"  NDA: {i}/{len(matching)} structures, {len(rows)} elements…",
                  file=sys.stderr)

    if not rows:
        raise DictionaryError("NDA returned no data elements")
    return rows, {
        "source": NDA_DICT_API,
        "source_sha256": None,
        "source_format": "nda_data_dictionary_api",
        "nda_structures_matched": len(matching),
        "nda_structures_fetched_now": fetched,
        "nda_structures_failed": failed,
        "nbdctools_data_releases_available": None,
    }


CATALOG_FILENAMES = ("NBDC_variable_catalog_full.xlsx", "NBDC_variable_catalog.xlsx")


def find_catalog(explicit: Optional[Path] = None) -> Optional[Path]:
    """Locate the NBDC variable catalog workbook without being told where it is.

    Order: an explicit path, then the skill's own data dir, then the usual places a
    download lands. Keeps `--from-xlsx` optional instead of hard-coding somebody's
    ~/Downloads into a command.
    """
    if explicit:
        p = Path(explicit).expanduser()
        return p if p.is_file() else None
    for base in (BUNDLED_DIR.parent, Path.cwd(), Path.home() / "Downloads",
                 Path.home() / "Desktop", Path.home()):
        for name in CATALOG_FILENAMES:
            cand = base / name
            if cand.is_file():
                return cand
    return None


def _parse_sheet_name(sheet: str) -> Optional[Tuple[str, str]]:
    """'ABCD 6.1' -> ('abcd', '6.1'). None for non-release sheets (legends etc.)."""
    m = re.match(r"^\s*(ABCD|HBCD)\s+([0-9]+(?:\.[0-9]+)?)\s*$", sheet, re.I)
    return (m.group(1).lower(), m.group(2)) if m else None


def xlsx_sheets(path: Path) -> List[Tuple[str, str, str]]:
    """Release sheets in a catalog workbook as (sheet_name, study, release)."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise DictionaryError(
            "reading an .xlsx catalog needs openpyxl — `pip install openpyxl`"
        ) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        out = []
        for name in wb.sheetnames:
            parsed = _parse_sheet_name(name)
            if parsed:
                out.append((name, parsed[0], parsed[1]))
        return out
    finally:
        wb.close()


def _rows_from_xlsx(path: Path, study: str, release: str, *,
                    sheet: Optional[str] = None,
                    progress: bool = False) -> Tuple[List[dict], dict]:
    """Read one release sheet of the NBDC variable catalog workbook.

    The catalog (NBDC_variable_catalog_full.xlsx) has one sheet per study+release
    with ~40 columns and 80-95k rows. Only the columns needed for verification and
    description are kept — a full copy would make the snapshot enormous for no
    gain.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise DictionaryError(
            "reading an .xlsx catalog needs openpyxl — `pip install openpyxl`"
        ) from exc

    wanted = set(KEEP_COLUMNS) | set(NAME_ALIAS_COLUMNS) | {
        "table_label", "url_table", "url_docs_score", "type_var",
    }

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        target = sheet
        if target is None:
            for name in wb.sheetnames:
                parsed = _parse_sheet_name(name)
                if parsed == (study, release):
                    target = name
                    break
        if target is None:
            have = ", ".join(f"{s}" for s, _, _ in xlsx_sheets(path)) or "none"
            raise DictionaryError(
                f"no sheet for {study} {release} in {path.name} "
                f"(release sheets: {have})"
            )

        ws = wb[target]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            raise DictionaryError(f"sheet {target!r} is empty")
        cols = {str(h).strip(): i for i, h in enumerate(header) if h}
        if "name" not in cols:
            raise DictionaryError(
                f"sheet {target!r} has no `name` column (saw: "
                f"{', '.join(list(cols)[:10])})"
            )
        keep = {c: i for c, i in cols.items() if c in wanted}

        rows: List[dict] = []
        for n, raw in enumerate(it, 1):
            name = raw[cols["name"]] if cols["name"] < len(raw) else None
            if name is None or not str(name).strip():
                continue
            rec: Dict[str, Any] = {}
            for col, idx in keep.items():
                if idx < len(raw):
                    val = raw[idx]
                    if val not in (None, ""):
                        rec[col] = val if isinstance(val, (str, int, float, bool)) \
                            else str(val)
            rows.append(_keep(rec))
            if progress and n % 20000 == 0:
                print(f"  {target}: {n} rows…", file=sys.stderr)
    finally:
        wb.close()

    if not rows:
        raise DictionaryError(f"sheet {target!r} produced no variables")
    return rows, {
        "source": str(path),
        "source_sha256": _sha256_file(path),
        "source_sheet": target,
        "source_format": "xlsx_variable_catalog",
        "nbdctools_data_releases_available": None,
    }


def build_snapshot(study: str, release: str, *, out_dir: Path = DEFAULT_DIR,
                   from_csv: Optional[Path] = None, from_rds: Optional[Path] = None,
                   from_xlsx: Optional[Path] = None, sheet: Optional[str] = None,
                   from_nda: bool = False, nda_prefix: Optional[str] = None,
                   minimal: bool = False, gzip_out: bool = False,
                   progress: bool = False) -> Path:
    """Write one study+release snapshot and return its path."""
    study = study.lower().strip()
    if study not in ("abcd", "hbcd"):
        raise DictionaryError(f"study must be 'abcd' or 'hbcd', got {study!r}")

    if from_nda:
        rows, prov = _rows_from_nda(study, cache_dir=out_dir, prefix=nda_prefix,
                                    progress=progress)
        resolved = release
    elif from_xlsx is not None:
        rows, prov = _rows_from_xlsx(Path(from_xlsx), study, release, sheet=sheet,
                                     progress=progress)
        resolved = release
    elif from_csv is not None:
        rows, prov = _rows_from_csv(Path(from_csv))
        resolved = release
    else:
        rows, prov = _rows_from_nbdctools(study, release, rds_path=from_rds,
                                          progress=progress)
        resolved = prov.pop("resolved_release")

    docs_note: Dict[str, Any] = {}
    if study == "abcd":
        docs = documented_releases(dir_=out_dir)
        known = docs.get("releases") or []
        docs_note = {
            "documentation_release_notes": release_citation(resolved),
            "documented_releases": known or None,
            "release_documented": (resolved in known) if known else None,
        }
        if known and resolved not in known:
            # Advisory, not fatal: the snapshot is real, but a release the public
            # documentation does not list is worth surfacing rather than burying.
            print(f"note: release {resolved!r} is not listed at "
                  f"{ABCD_RELEASE_NOTES_INDEX} (documented: {', '.join(known)})",
                  file=sys.stderr)

    snapshot = {
        "study": study,
        "release": resolved,
        "variable_count": len(rows),
        "columns": sorted({k for r in rows for k in r}),
        "provenance": {
            "retrieved_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "method": ("nda_api" if from_nda
                       else "xlsx_catalog" if from_xlsx
                       else "csv_export" if from_csv else "nbdctools"),
            "tool_version": _nbdctools_version(),
            **docs_note,
            **prov,
        },
        "variables": rows,
    }
    if minimal:
        snapshot["variables"] = [
            {k: v for k, v in row.items() if k in MINIMAL_COLUMNS and v not in (None, "")}
            for row in snapshot["variables"]
        ]
        snapshot["columns"] = sorted({k for r in snapshot["variables"] for k in r})
        snapshot["provenance"]["projection"] = "minimal"
        snapshot["provenance"]["minimal_columns"] = list(MINIMAL_COLUMNS)

    out_dir.mkdir(parents=True, exist_ok=True)
    if gzip_out:
        import gzip as _gzip

        out = out_dir / f"dd-{study}-{resolved}.json.gz"
        out.write_bytes(_gzip.compress(
            json.dumps(snapshot, separators=(",", ":")).encode("utf-8"), 9))
    else:
        out = out_dir / f"dd-{study}-{resolved}.json"
        out.write_text(json.dumps(snapshot, indent=1, sort_keys=False))
    return out


def _nbdctools_version() -> Optional[str]:
    try:
        import nbdctools  # type: ignore

        return getattr(nbdctools, "__version__", None)
    except Exception:
        return None


def documented_releases(*, dir_: Path = DEFAULT_DIR, timeout: int = 20,
                        refresh: bool = False) -> dict:
    """Releases documented on docs.abcdstudy.org, cached locally.

    Advisory only — a release missing here is reported, never rejected. The docs
    site is public prose, so it can lag or restructure; it must not be able to
    block a run whose dictionary snapshot is real.
    """
    cache = dir_ / "abcd_documented_releases.json"
    if cache.is_file() and not refresh:
        try:
            return json.loads(cache.read_text())
        except Exception:
            pass

    import urllib.request

    doc = {"releases": [], "source": ABCD_RELEASE_NOTES_INDEX,
           "retrieved_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
           "error": None}
    try:
        req = urllib.request.Request(
            ABCD_RELEASE_NOTES_INDEX, headers={"Accept": "text/html"}
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            html_text = resp.read().decode("utf-8", "replace")
        found = sorted(
            {m.replace("_", ".")
             for m in re.findall(r"release_notes/(\d+_\d+)\.html", html_text)},
            key=_release_sort_key,
        )
        doc["releases"] = found
    except Exception as exc:                              # advisory: never fatal
        doc["error"] = str(exc)

    try:
        dir_.mkdir(parents=True, exist_ok=True)
        cache.write_text(json.dumps(doc, indent=1))
    except Exception:
        pass
    return doc


def release_citation(release: str) -> str:
    """Citable release-notes URL for an ABCD release ('6.1' -> …/6_1.html)."""
    return (f"https://docs.abcdstudy.org/latest/documentation/release_notes/"
            f"{str(release).replace('.', '_')}.html")


def available_releases(study: str, *, progress: bool = False) -> List[str]:
    """Every release the bundle offers for `study` (needs nbdctools)."""
    _, prov = _rows_from_nbdctools(study, "latest", rds_path=None, progress=progress)
    return prov.get("nbdctools_data_releases_available") or []


# --------------------------------------------------------------------------- #
# snapshot loading + resolution
# --------------------------------------------------------------------------- #

@dataclass
class Match:
    """One dictionary hit for a string found in a paper."""

    name: str
    study: str
    release: str
    method: str                     # exact_name | normalized_name | label | description
    score: float                    # 1.0 for exact, lower for text matches
    row: dict = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            "variable": self.name,
            "study": self.study,
            "dd_release": self.release,
            "match_method": self.method,
            "match_score": round(self.score, 3),
            "label": self.row.get("label"),
            # Canonical names, resolved through FIELD_ALIASES so a bundle export
            # and a portal CSV produce the same keys.
            "nda_or_nbdc_table": canonical_field(self.row, "nda_or_nbdc_table"),
            "nbdc_domain": canonical_field(self.row, "nbdc_domain"),
            "nbdc_sub_domain": canonical_field(self.row, "nbdc_sub_domain"),
            "nbdc_table": self.row.get("table_name"),
            "table_label": self.row.get("table_label"),
            "name_nda": self.row.get("name_nda"),
            "name_deap": self.row.get("name_deap"),
            "url_table": self.row.get("url_table"),
        }


class Dictionary:
    """One or more release snapshots, queried together.

    Holding several releases at once is deliberate: papers cite variables from
    whatever release they analysed, and names change between releases. Resolving
    across all loaded releases lets us say "present in 5.1 and 6.0, absent in
    6.1" instead of a bare miss.
    """

    def __init__(self, snapshots: List[dict]):
        if not snapshots:
            raise DictionaryError(
                "No dictionary snapshots loaded. Run: python -m "
                "scripts.abcd_dictionary build --study abcd --release latest"
            )
        self.snapshots = snapshots
        self._by_name: Dict[str, List[Match]] = {}
        self._by_label: Dict[str, List[Match]] = {}
        self._rows: List[Tuple[dict, str, str]] = []
        for snap in snapshots:
            study, rel = snap["study"], snap["release"]
            for row in snap["variables"]:
                name = str(row.get("name", ""))
                if not name:
                    continue
                self._rows.append((row, study, rel))
                self._by_name.setdefault(norm_name(name), []).append(
                    Match(name, study, rel, "exact_name", 1.0, row)
                )
                # Alternate namings (NDA element name, DEAP name, REDCap, short,
                # Stata). A paper citing any of them is citing this variable, and
                # the match method records which naming it used.
                for col, method in NAME_ALIAS_COLUMNS.items():
                    alt = str(row.get(col) or "").strip()
                    if alt and norm_name(alt) != norm_name(name):
                        self._by_name.setdefault(norm_name(alt), []).append(
                            Match(name, study, rel, method, 0.98, row)
                        )
                for col in ("label", "description"):
                    txt = norm_text(str(row.get(col) or ""))
                    if len(txt) >= 8:
                        self._by_label.setdefault(txt, []).append(
                            Match(name, study, rel, col, 0.9, row)
                        )

    # -- construction ------------------------------------------------------- #

    @classmethod
    def load(cls, *, study: Optional[str] = None, releases: Optional[Iterable[str]] = None,
             dir_: Path = DEFAULT_DIR) -> "Dictionary":
        snaps = load_all_snapshots(dir_=dir_)
        if study:
            snaps = [s for s in snaps if s["study"] == study.lower()]
        if releases:
            want = {str(r) for r in releases}
            snaps = [s for s in snaps if s["release"] in want]
        return cls(snaps)

    @property
    def provenance(self) -> List[dict]:
        """What every consulted snapshot was — goes into the run's provenance."""
        return [
            {
                "study": s["study"],
                "dd_release": s["release"],
                "variable_count": s["variable_count"],
                **{k: v for k, v in s.get("provenance", {}).items()
                   if k in ("retrieved_at", "method", "source", "source_sha256",
                            "tool_version", "documentation_release_notes",
                            "release_documented", "csv_header_mapping")},
            }
            for s in self.snapshots
        ]

    # -- queries ------------------------------------------------------------ #

    def resolve(self, candidate: str, *, allow_label_match: bool = True) -> List[Match]:
        """All dictionary hits for `candidate`, best first, one per (name, release).

        Exact name match is tried first. Label/description matching is a fallback
        for papers that name a measure in prose ("NIH Toolbox Flanker score")
        rather than by variable id, and requires the FULL label to match — no
        substring or fuzzy scoring, because a partial label match is not evidence
        that a specific variable was used.
        """
        key = norm_name(candidate)
        hits = list(self._by_name.get(key, []))

        if not hits and "__" in key:
            # Some papers write single underscores where the dictionary uses
            # double (table__column). Try the double-underscore reading too.
            hits = list(self._by_name.get(key.replace("__", "_"), []))
        if not hits and "_" in key:
            for alt, matches in self._by_name.items():
                if alt.replace("__", "_") == key.replace("__", "_"):
                    hits.extend(m for m in matches)

        if hits:
            for h in hits:
                if h.method == "exact_name" and norm_name(h.name) != key:
                    h.method, h.score = "normalized_name", 0.95
            return _dedupe(hits)

        if allow_label_match:
            lab = norm_text(candidate)
            if len(lab) >= 8:
                return _dedupe(self._by_label.get(lab, []))
        return []

    def releases_for(self, name: str) -> List[str]:
        """Releases whose dictionary contains `name` — the rename detector."""
        return sorted(
            {m.release for m in self._by_name.get(norm_name(name), [])},
            key=_release_sort_key,
        )

    def search(self, needle: str, limit: int = 20) -> List[Match]:
        """Substring search over names and labels. For humans exploring, NOT for
        verification — `resolve()` is what decides whether a claim holds."""
        n = norm_text(needle)
        out: List[Match] = []
        for row, study, rel in self._rows:
            hay = " ".join(
                str(row.get(c) or "") for c in ("name", "label", "description")
            ).lower()
            if n in hay:
                out.append(Match(str(row["name"]), study, rel, "search", 0.5, row))
                if len(out) >= limit:
                    break
        return out


def _dedupe(matches: List[Match]) -> List[Match]:
    best: Dict[Tuple[str, str, str], Match] = {}
    for m in matches:
        k = (m.study, m.release, norm_name(m.name))
        if k not in best or m.score > best[k].score:
            best[k] = m
    return sorted(best.values(), key=lambda m: (-m.score, _release_sort_key(m.release)))


def load_snapshot(path: Path) -> dict:
    path = Path(path)
    if path.suffix == ".gz":
        import gzip

        snap = json.loads(gzip.decompress(path.read_bytes()).decode("utf-8"))
    else:
        snap = json.loads(path.read_text())
    for key in ("study", "release", "variables"):
        if key not in snap:
            raise DictionaryError(f"{path} is not a dictionary snapshot (missing {key!r})")
    return snap


def load_all_snapshots(*, dir_: Path = DEFAULT_DIR,
                       include_bundled: bool = True) -> List[dict]:
    """Snapshots from `dir_`, falling back to the ones bundled with the skill.

    A locally built snapshot WINS over a bundled one for the same study+release:
    the bundle is a convenience so the skill works out of the box, not a ceiling.
    Rebuild from the catalog workbook whenever a new release lands.
    """
    found: Dict[Tuple[str, str], dict] = {}
    sources: List[Path] = []
    if include_bundled and BUNDLED_DIR.is_dir():
        sources.append(BUNDLED_DIR)
    if dir_.is_dir():
        sources.append(dir_)          # later wins
    for src in sources:
        for pattern in ("dd-*.json", "dd-*.json.gz"):
            for pth in sorted(src.glob(pattern)):
                try:
                    snap = load_snapshot(pth)
                except Exception:
                    continue
                snap.setdefault("provenance", {})["snapshot_path"] = str(pth)
                snap["provenance"]["bundled_with_skill"] = (src == BUNDLED_DIR)
                found[(snap["study"], snap["release"])] = snap
    return [found[k] for k in sorted(found, key=lambda k: (k[0], _release_sort_key(k[1])))]


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def _cli(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=(__doc__ or "").splitlines()[0])
    sub = ap.add_subparsers(dest="cmd", required=True)

    b = sub.add_parser("build", help="materialise a dictionary snapshot")
    b.add_argument("--study", default="abcd", choices=["abcd", "hbcd"])
    b.add_argument("--release", default="latest",
                   help="release id (e.g. 6.1) or 'latest'")
    b.add_argument("--all-releases", action="store_true",
                   help="snapshot every release the bundle offers")
    b.add_argument("--from-csv", type=Path, help="build from your own CSV/TSV export")
    b.add_argument("--from-xlsx", type=Path,
                   help="build from an NBDC variable catalog workbook "
                        "(NBDC_variable_catalog_full.xlsx)")
    b.add_argument("--sheet", help="explicit sheet name in the workbook")
    b.add_argument("--from-nda", action="store_true",
                   help="build from NDA structure definitions (public) — the source "
                        "for pre-6.0 releases (4.x / 5.x), which the NBDC catalog "
                        "and NBDCtools do not cover")
    b.add_argument("--nda-prefix", help="structure prefix to match (default <study>_)")
    b.add_argument("--all-sheets", action="store_true",
                   help="with --from-xlsx: build every release sheet in the workbook")
    b.add_argument("--from-rds", type=Path, help="use an already-downloaded lst_dds.rds")
    b.add_argument("--dir", type=Path, default=DEFAULT_DIR)
    b.add_argument("--minimal", action="store_true",
                   help="keep only the columns the ABCD mode reads (much smaller)")
    b.add_argument("--gzip", action="store_true", dest="gzip_out",
                   help="write dd-<study>-<release>.json.gz")
    b.add_argument("--progress", action="store_true")

    i = sub.add_parser("info", help="list local snapshots")
    i.add_argument("--dir", type=Path, default=DEFAULT_DIR)

    r2 = sub.add_parser("releases",
                        help="releases documented at docs.abcdstudy.org (public)")
    r2.add_argument("--dir", type=Path, default=DEFAULT_DIR)
    r2.add_argument("--refresh", action="store_true")

    l = sub.add_parser("lookup", help="resolve a variable name / label")
    l.add_argument("candidate")
    l.add_argument("--study")
    l.add_argument("--dir", type=Path, default=DEFAULT_DIR)

    s = sub.add_parser("search", help="substring search (exploration only)")
    s.add_argument("needle")
    s.add_argument("--limit", type=int, default=20)
    s.add_argument("--dir", type=Path, default=DEFAULT_DIR)

    a = ap.parse_args(argv)

    # `build --from-xlsx` with no path: find the workbook.
    if getattr(a, "cmd", None) == "build" and getattr(a, "from_xlsx", None) is None \
            and not a.from_csv and not a.from_rds and not a.from_nda:
        found = find_catalog()
        if found is not None:
            a.from_xlsx = found
            print(f"using catalog workbook: {found}", file=sys.stderr)

    try:
        if a.cmd == "build":
            if a.from_xlsx and a.all_sheets:
                sheets = xlsx_sheets(Path(a.from_xlsx))
                if not sheets:
                    print(f"no release sheets found in {a.from_xlsx}", file=sys.stderr)
                    return 1
                for sheet_name, study, rel in sheets:
                    out = build_snapshot(study, rel, out_dir=a.dir,
                                         from_xlsx=a.from_xlsx, sheet=sheet_name,
                                         minimal=a.minimal, gzip_out=a.gzip_out,
                                         progress=a.progress)
                    snap = load_snapshot(out)
                    print(f"wrote {out}  ({snap['variable_count']} variables)")
                return 0

            targets = [a.release]
            if a.all_releases:
                targets = available_releases(a.study, progress=a.progress)
                print(f"releases: {', '.join(targets)}", file=sys.stderr)
            for rel in targets:
                out = build_snapshot(a.study, rel, out_dir=a.dir, from_csv=a.from_csv,
                                     from_rds=a.from_rds, from_xlsx=a.from_xlsx,
                                     sheet=a.sheet, from_nda=a.from_nda,
                                     nda_prefix=a.nda_prefix, minimal=a.minimal,
                                     gzip_out=a.gzip_out, progress=a.progress)
                snap = load_snapshot(out)
                print(f"wrote {out}  ({snap['variable_count']} variables)")
            return 0

        if a.cmd == "releases":
            docs = documented_releases(dir_=a.dir, refresh=a.refresh)
            if docs.get("error"):
                print(f"could not read {docs['source']}: {docs['error']}",
                      file=sys.stderr)
                return 1
            for rel in docs["releases"]:
                print(f"{rel:6} {release_citation(rel)}")
            print(f"(retrieved {docs['retrieved_at']})", file=sys.stderr)
            return 0

        if a.cmd == "info":
            snaps = load_all_snapshots(dir_=a.dir)
            if not snaps:
                print(f"no snapshots in {a.dir}", file=sys.stderr)
                return 1
            for s_ in snaps:
                p = s_.get("provenance", {})
                print(f"{s_['study']:5} {s_['release']:6} {s_['variable_count']:>7} vars"
                      f"  via {p.get('method')}  {p.get('retrieved_at')}")
            return 0

        d = Dictionary.load(study=getattr(a, "study", None), dir_=a.dir)
        if a.cmd == "lookup":
            hits = d.resolve(a.candidate)
            if not hits:
                print(f"UNVERIFIED: {a.candidate!r} is not in any loaded dictionary")
                return 2
            for m in hits:
                print(json.dumps(m.to_dict()))
            rel = d.releases_for(hits[0].name)
            if rel:
                print(f"present in releases: {', '.join(rel)}", file=sys.stderr)
            return 0

        if a.cmd == "search":
            for m in d.search(a.needle, limit=a.limit):
                print(f"{m.name:44} {str(m.row.get('label'))[:60]}")
            return 0
    except DictionaryError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(_cli())
